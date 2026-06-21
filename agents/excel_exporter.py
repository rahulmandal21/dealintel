import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os

def export_deal_to_excel(
    target: str,
    acquirer: str,
    revenue_by_year: dict,
    net_income_by_year: dict,
    dcf_low: float,
    dcf_high: float,
    comps_low: float,
    comps_high: float,
    deal_premium: float = 25.0,
    output_path: str = "dealintel_output.xlsx"
):
    wb = openpyxl.Workbook()

    # ── Styles ───────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(cell, text):
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    def style_cell(cell, value, fill=None, bold_text=False, number_format=None):
        cell.value = value
        if fill:
            cell.fill = fill
        if bold_text:
            cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = thin
        if number_format:
            cell.number_format = number_format

    # ══════════════════════════════════════════════════════════════
    # SHEET 1: DCF Model
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "DCF Model"
    ws1.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 18

    # Title
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = f"DCF Valuation — {acquirer} acquiring {target}"
    title_cell.fill = header_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = center
    ws1.row_dimensions[1].height = 30

    # Assumptions
    ws1["A3"] = "DCF Assumptions"
    ws1["A3"].font = Font(bold=True, size=12)
    assumptions = [
        ("WACC", "10.0%"),
        ("Terminal Growth Rate", "3.0%"),
        ("Deal Premium", f"{deal_premium}%"),
        ("Projection Years", "5"),
    ]
    for i, (label, val) in enumerate(assumptions, start=4):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = val
        ws1[f"A{i}"].font = bold
        ws1[f"B{i}"].alignment = center

    # Revenue table
    row = 10
    style_header(ws1[f"A{row}"], "Metric")
    years = sorted(revenue_by_year.keys())[-5:]
    for j, yr in enumerate(years):
        style_header(ws1[f"{get_column_letter(j+2)}{row}"], yr)

    metrics = [
        ("Revenue ($)", revenue_by_year),
        ("Net Income ($)", net_income_by_year),
    ]
    for metric_name, data in metrics:
        row += 1
        ws1[f"A{row}"].value = metric_name
        ws1[f"A{row}"].font = bold
        ws1[f"A{row}"].border = thin
        for j, yr in enumerate(years):
            val = data.get(yr, 0)
            style_cell(
                ws1[f"{get_column_letter(j+2)}{row}"],
                val, number_format='#,##0'
            )

    # Valuation summary
    row += 3
    style_header(ws1[f"A{row}"], "Valuation Summary")
    ws1.merge_cells(f"A{row}:F{row}")
    ws1[f"A{row}"].alignment = center

    row += 1
    for label, low, high, fill in [
        ("DCF Valuation Range ($B)", dcf_low, dcf_high, blue_fill),
        ("Comps Valuation Range ($B)", comps_low, comps_high, green_fill),
    ]:
        ws1[f"A{row}"] = label
        ws1[f"A{row}"].font = bold
        ws1[f"A{row}"].border = thin
        style_cell(ws1[f"B{row}"], f"${low}B - ${high}B", fill=fill)
        ws1.merge_cells(f"B{row}:F{row}")
        row += 1

    implied_low = round(comps_low * (1 + deal_premium/100), 1)
    implied_high = round(comps_high * (1 + deal_premium/100), 1)
    ws1[f"A{row}"] = f"Implied Deal Value ({deal_premium}% premium)"
    ws1[f"A{row}"].font = Font(bold=True, color="FFFFFF")
    ws1[f"A{row}"].fill = header_fill
    ws1[f"A{row}"].border = thin
    style_cell(ws1[f"B{row}"], f"${implied_low}B - ${implied_high}B",
               fill=header_fill)
    ws1[f"B{row}"].font = Font(bold=True, color="FFFFFF")
    ws1.merge_cells(f"B{row}:F{row}")

    # ══════════════════════════════════════════════════════════════
    # SHEET 2: Sensitivity Table
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Sensitivity Analysis")
    ws2.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F","G"]:
        ws2.column_dimensions[col].width = 16

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "Sensitivity Analysis — EV/EBITDA Multiple vs Deal Premium"
    ws2["A1"].fill = header_fill
    ws2["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 28

    ws2["A3"] = "EV/EBITDA \\ Premium"
    ws2["A3"].font = bold
    ws2["A3"].border = thin

    premiums = [10, 15, 20, 25, 30, 35]
    multiples = [4, 6, 8, 10, 12, 14]
    base_ebitda = (comps_low + comps_high) / 2 / 6

    for j, p in enumerate(premiums):
        cell = ws2[f"{get_column_letter(j+2)}3"]
        style_header(cell, f"{p}%")

    for i, mult in enumerate(multiples):
        row_num = i + 4
        ws2[f"A{row_num}"].value = f"{mult}x"
        ws2[f"A{row_num}"].font = bold
        ws2[f"A{row_num}"].border = thin
        ws2[f"A{row_num}"].alignment = center
        for j, prem in enumerate(premiums):
            implied = round(base_ebitda * mult * (1 + prem/100), 1)
            cell = ws2[f"{get_column_letter(j+2)}{row_num}"]
            cell.value = f"${implied}B"
            cell.alignment = center
            cell.border = thin
            # Color code
            if implied < comps_low:
                cell.fill = red_fill
            elif implied > comps_high * 1.2:
                cell.fill = green_fill
            else:
                cell.fill = yellow_fill

    # Legend
    ws2["A12"] = "🟢 Above range  🟡 In range  🔴 Below range"
    ws2["A12"].font = Font(italic=True)

    # ══════════════════════════════════════════════════════════════
    # SHEET 3: Football Field
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Football Field")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20

    ws3.merge_cells("A1:C1")
    ws3["A1"].value = f"Football Field Chart — {target} Valuation"
    ws3["A1"].fill = header_fill
    ws3["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws3["A1"].alignment = center

    style_header(ws3["A3"], "Methodology")
    style_header(ws3["B3"], "Low ($B)")
    style_header(ws3["C3"], "High ($B)")

    methods = [
        ("DCF Valuation", dcf_low, dcf_high, blue_fill),
        ("Comparable Transactions", comps_low, comps_high, green_fill),
        (f"Implied Deal Value ({deal_premium}% premium)",
         implied_low, implied_high, yellow_fill),
    ]
    for i, (name, low, high, fill) in enumerate(methods, start=4):
        ws3[f"A{i}"].value = name
        ws3[f"A{i}"].font = bold
        ws3[f"A{i}"].border = thin
        style_cell(ws3[f"B{i}"], low, fill=fill, number_format='#,##0.0')
        style_cell(ws3[f"C{i}"], high, fill=fill, number_format='#,##0.0')

    # Bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Football Field Valuation"
    chart.y_axis.title = "Valuation ($B)"
    chart.x_axis.title = "Method"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data_ref = Reference(ws3, min_col=2, max_col=3, min_row=3, max_row=6)
    cats = Reference(ws3, min_col=1, min_row=4, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws3.add_chart(chart, "E3")

    # Save
    wb.save(output_path)
    print(f"✅ Excel file saved: {output_path}")
    return output_path


if __name__ == "__main__":
    # Test with sample data
    export_deal_to_excel(
        target="NVS (Novartis)",
        acquirer="HINDALCO",
        revenue_by_year={"2021": 51626000000, "2022": 50545000000,
                         "2023": 45440000000, "2024": 47445000000},
        net_income_by_year={"2021": 11779000000, "2022": 6956000000,
                            "2023": 8572000000, "2024": 9200000000},
        dcf_low=100,
        dcf_high=150,
        comps_low=60,
        comps_high=120,
        deal_premium=25.0,
        output_path="dealintel_output.xlsx"
    )